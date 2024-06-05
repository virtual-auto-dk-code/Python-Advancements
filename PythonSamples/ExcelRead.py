#Program to test Excel Reading in Selenium
import openpyxl

workbook=openpyxl.load_workbook("QueryResult.xlsx")
sheet=workbook["Sheet1"]
totalrows=sheet.max_row
totalcols=sheet.max_column
print("Total no of rows in sheet1 are: ", totalrows , "Total no of rows in sheet1 are: ", totalcols)

#Display single cell value
print(sheet.cell(row=2,column=1).value)

#Display all the contents of sheet
for row in range(1,totalrows+1):
    for col in range(1,totalcols+1):
        print(sheet.cell(row=row,column=col).value,end=" ")
    print()